from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import openpyxl

driver = webdriver.Chrome()
url = "http://exam.smvec.ac.in/exam_result_ug_regular_mayjune_2024/"
driver.get(url)
time.sleep(5)




def read_captcha():
    captcha =  driver.find_element("xpath","/html/body/div[3]/div[1]/div/div/div[2]/div[2]/div[2]/div/div/div/div/form/div[3]/div/span[1]")
    return captcha.text

def get_result(regno_value,dob_value):
    regno = driver.find_element("xpath","/html/body/div[3]/div[1]/div/div/div[2]/div[2]/div[2]/div/div/div/div/form/div[1]/div/input")
    regno.send_keys(regno_value)

    dob = driver.find_element("xpath","/html/body/div[3]/div[1]/div/div/div[2]/div[2]/div[2]/div/div/div/div/form/div[2]/div/input")
    dob.send_keys(dob_value)

    captcha = driver.find_element("xpath","/html/body/div[3]/div[1]/div/div/div[2]/div[2]/div[2]/div/div/div/div/form/div[3]/div/input")
    captcha.send_keys(read_captcha())
    
    submit = driver.find_element("xpath","/html/body/div[3]/div[1]/div/div/div[2]/div[2]/div[2]/div/div/div/div/form/div[4]/div/button[1]")
    submit.click()
    
def convert_date(date):
    lst = date.split(".")
    return lst[0]+"/"+lst[1]+"/"+lst[2]
    
    
def read_result():
    result = dict()
    table = driver.find_element(By.XPATH, "/html/body/div[3]/div[1]/div/div/div[3]/div[2]/div/div/div/div/div/div/div[6]/table")
    rows = table.find_elements(By.TAG_NAME, "tr")
    for row in rows[1:]:  
        columns = row.find_elements(By.TAG_NAME, "td")
        if len(columns) == 6:  
            subject = columns[2].text
            marks = columns[3].text
            result[subject] = marks
    sgpa_web = driver.find_element("xpath","/html/body/div[3]/div[1]/div/div/div[3]/div[2]/div/div/div/div/div/div/div[7]")
    sgpa=sgpa_web.text
    
    reset = driver.find_element("xpath","/html/body/div[3]/div[1]/div/div/div[2]/div[2]/div[2]/div/div/div/div/form/div[4]/div/button[2]")
    reset.click()
    
    return (result,sgpa)



workbook = openpyxl.load_workbook('21-25 IT C.xlsx')
worksheet = workbook.active

idx=2
num_rows = worksheet.max_row

print(num_rows)

while num_rows>=idx:
    reg_no=worksheet.cell(row=idx,column=2).value
    dob = worksheet.cell(row=idx,column=3).value
    
    print(reg_no,dob) 

    try:
        time.sleep(0.5)
        get_result(reg_no,convert_date(dob))
        time.sleep(0.5)
        result,sgpa = read_result()
        print(result,sgpa)
    except Exception as e:
        idx += 1
        continue
    
    col_mark=4
    for mark in result:
        print(mark , result[mark])
        worksheet.cell(row=1,column=col_mark).value = mark
        worksheet.cell(row=idx,column=col_mark).value = result[mark]
        col_mark+=1
    worksheet.cell(row=1,column=col_mark).value="SGPA"
    worksheet.cell(row=idx,column=col_mark).value=sgpa[5:]
    idx+=1
    print("\n\n")
    time.sleep(0.5)
        

workbook.save("Results/21-25 IT C Results.xlsx")
driver.close()
